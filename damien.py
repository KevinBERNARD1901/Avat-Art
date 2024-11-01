# Ceci est la page de Daminus

from pykinect2 import PyKinectV2
from pykinect2 import PyKinectRuntime
import pygame

# Initialisation de la kinect
kinect = PyKinectRuntime.PyKinectRuntime(PyKinectV2.FrameSourceTypes_Body | PyKinectV2.FrameSourceTypes_Color)

# Initialisation de pygame
pygame.init()
ecran_width, ecran_height = 960, 540
ecran = pygame.display.set_mode((ecran_width, ecran_height))

# Fonction pour dessiner un squelette
def draw_body(ecran, joints):
    for joint in joints:
        position = joints[joint].Position
        x, y, z = position.x, position.y, position.z
        if z > 0:
            x, y = int(x * 100 + ecran_width // 2), int(-y * 100 + ecran_height // 2)
            pygame.draw.circle(ecran, (255, 0, 0), (x, y), 5)

# Ouverture du fichier (avec l'aide de monsieur GPT parceque vraiment trop à la bourre)
import openpyxl
chemin_fichier_excel = "" # A REMPLIR !!!
classeur = openpyxl.load_workbook(chemin_fichier_excel)
feuille = classeur["Acquisition"]
ligne_acquisiton = 3 # A implémenter de 3 à chaque acquisition
Nb_acquisition_faite = 0 # A implémenter de 1 à chaque acquisition


# Boucle principale
running = True
while running:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
    
    # Récupération des données de la kinect
    if kinect.has_new_body_frame():
        bodies = kinect.get_last_body_frame()
        if bodies != None:
            for i in range(0, kinect.max_body_count):
                body = bodies.bodies[i]
                if not body.is_tracked:
                    continue
                joints = body.joints

                # !!!!!! Manque une condition avec la main gauche fermée !!!!!!

                # Récupération de la position de toutes les jointures
                for i in range(len(joints)):

                    #Récupération
                    joint = joints[i] # Fonctionne ???
                    position = joint.Position
                    x, y, z = position.x, position.y, position.z

                    # Remplissage excel
                    feuille.cell(row=ligne_acquisiton,column=i+3, value=x)
                    feuille.cell(row=ligne_acquisiton+1,column=i+3, value=y)
                    feuille.cell(row=ligne_acquisiton+2,column=i+3, value=z)

                    #Mise à jour des variable de navigation dans l'excel
                    ligne_acquisiton += 3
                    Nb_acquisition_faite += 1

                    if Nb_acquisition_faite > 100:
                        running = False

                    draw_body(ecran, joints)

# Enregistrement
classeur.save(chemin_fichier_excel)
print("Modification enregistrée avec succès !")